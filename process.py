from math import sqrt
from openpyxl import Workbook,load_workbook
import numpy as np

class Point3d():
    _x:float
    _y:float
    _z:float

    def __init__(self,X:float, Y:float, Z:float) -> None:
        self._x = X
        self._y = Y
        self._z = Z
        pass

    def __str__(self) -> str:
        return f'{self._x},{self._y},{self._z}'
    
    def getX(self) -> float:
        return self._x
    
    def getY(self) -> float:
        return self._y
    
    def getZ(self) -> float:
        return self._z

    def distance3dTo(self, point:'Point3d'):
        return sqrt((point.getX() - self._x)**2 + (point.getY() - self._y)**2 + (point.getZ() - self._z)**2)

    def distance2dTo(self,point:'Point3d'):
        return sqrt((point.getX() - self._x)**2 + (point.getY() - self._y)**2)
    
    def move(self, vector:'Vector3d') -> 'Point3d':
        x = self._x + vector.getX()
        y = self._y + vector.getY() 
        z = self._z + vector.getZ()
        return Point3d(x,y,z)
    
    def to_numpy(self) -> np.ndarray:
        return np.array([self._x,self._y,self._z])
    
    @staticmethod
    def getPointsInput(path:str) -> list['Point3d']:
        """
        Get a list of points from txt input
        """
        points = []
    
        with open(path) as f:
            lines = f.readlines()
            for line in lines:
                info = line.split("@")
                points.append(Point3d(float(info[0]),float(info[1]),float(info[2])))

        return points

class Vector3d(Point3d):
    __length :float
    
    def __init__(self, pointA:Point3d = Point3d(0,0,0), pointB:Point3d = Point3d(0,0,0),*, X=0,Y=0,Z=0) -> None:
        if (X != 0 or Y != 0 or Z != 0):
            super().__init__(X=X,Y=Y,Z=Z)
        else:
            super().__init__(
            X=pointB.getX() - pointA.getX(),
            Y=pointB.getY() - pointA.getY(),
            Z=pointB.getZ() - pointA.getZ())
            # self.__length = pointA.distance3dTo(pointB)
        
        self.__length = sqrt(pow(self._x,2) + pow(self._y,2) + pow(self._z,2))

    def __str__(self) -> str:
        return super().__str__()
    
    def getX(self) -> float:
        return self._x
    
    def getY(self) -> float:
        return self._y
    
    def getZ(self) -> float:
        return self._z
    
    def move(self,point:Point3d) -> Point3d:
        """
        Move a point by this vector
        """
        # print(f'{self._x} {self._y} {self._z}')
        x = self._x + point.getX()
        y = self._y + point.getY() 
        z = self._z + point.getZ()
        return Point3d(x,y,z)

    def getLength(self) -> float:
        return self.__length
    
    def getNormal(self) -> 'Vector3d':
        # print(f'{self._x} {self._y} {self._z}')
        x = self._x / self.__length
        z = self._z / self.__length
        y = self._y / self.__length
        return Vector3d(X=x,Y=y,Z=z)

    def multiplyBy(self, r:float) -> 'Vector3d':
        x = self._x * r
        y = self._y * r
        z = self._y * r
        return Vector3d(X=x,Y=y,Z=z)

    def dotProduct(self, vector:'Vector3d') -> float:
        return (self._x * vector.getX() + self._y * vector.getY() + self._z * vector.getZ())

class Utils():
    @staticmethod
    def getPath(path:str) -> str:
        """
        Return full path of parent directory of the given file
        """
        return path[:-len(path.split('/')[-1])]

    @staticmethod
    def txt2xlsx(input:str, output:str, separator:str = ","): #convert text file to xlsx
        
        savePath = Utils.getPath(input) + output
        print(savePath)
        wb = Workbook()
        ws = wb.active

        with open(input,'r') as inputf:
            lines = inputf.readlines()
            row = 1
            for line in lines:
                info = line.split(separator)
                for col in range(1,len(info)+1):
                    try:
                        ws.cell(row,col).value = float(info[col-1])
                    except ValueError:
                        ws.cell(row,col).value = info[col-1]
                row += 1

        wb.save(savePath)
    
    @staticmethod
    def xlsx2txt(path:str, output:str, sizeCol:int, sizeRow:int) -> None:
        """
        This function turn any xlsx to txt
        """
        wb = load_workbook(path)
        ws = wb.active
        
        output = Utils.getPath(path) + output
        for r in range(sizeRow):
            line = ""
            try:
                for c in range(sizeCol):
                    line += str(round(ws.cell(r+1,c+1).value,2)) + ","
                with open(output,'a') as f:
                    f.write(line[:-1]+"\n")
            except TypeError:
                continue
    @staticmethod
    def sortPoint(points:list, start:Point3d = None  ) -> list['Point3d']:
        """
        Take a list of points and sort it with smallest distance from a point to the next
        """
        if start == None: 
            start = points[0]
            points.pop(0)
        
        newPoints:list = [start]
        while len(points) > 0:
            mindist:float = np.Inf
            mindex = 0

            for i,point in enumerate(points):
                dist = start.distance2dTo(point)
                if dist < mindist:
                    mindist = dist
                    mindex = i

            start = points[mindex]
            newPoints.append(start)
            points.pop(mindex)
        return newPoints   
    

if __name__ == "__main__":
    pps = []
    with open("C:\\Users\\trong\\OneDrive\\Máy tính\\test.txt", "r") as file:
        lines = file.readlines()
        for line in lines:
            line = line.split(",")
            pps.append(Point3d(float(line[1]),float(line[2]),float(line[3])))
    
    sorted = Utils.sortPoint(pps)
    # print(sorted)

    with open("cotbenphai.csv", "w") as file:
        for i,p in enumerate(sorted):
            if i == 0:
                dist=0
            else:
                dist = p.distance3dTo(sorted[i-1])
            content = f'{p.getZ()},{dist}\n'
            file.write(content)