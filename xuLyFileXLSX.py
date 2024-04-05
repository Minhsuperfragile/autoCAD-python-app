from openpyxl import load_workbook, Workbook
from math import sqrt
from sys import argv as cmdArgs

class Point():
    x:float
    y:float
    z:float

    def __init__(self, x:float, y:float, z:float) -> None:
        self.x = x
        self.y = y
        self.z = z

    @staticmethod
    def distance(A, B) -> float:
        return sqrt((B.x-A.x)**2 + (B.y-A.y)**2 + (B.z-A.z)**2)

def getPath(path:str) -> str:
    return path[:-len(path.split('\\')[-1])]

def processXLSX(path:str, maxRow:int) -> None:
    wb =  load_workbook(path)
    kcl_list = [0] 
    kccd_list = [0]
    ws = wb.active

    filename = path.split('//')[-1]
    textfile = open(filename+'.txt',mode='w')

    for i in range(2,maxRow+1):       
        kcl_list.append(ws.cell(i-1,6).value)
        kccd_list.append(float(kccd_list[i-2]) + float(kcl_list[i-1]))
            
        line:str = str(round(kccd_list[i-2],2)) + '@' + str(round(kcl_list[i-2],2)) + '@' + str(round(float(ws.cell(i-1,4).value.replace(",",".")),2)) + '\n'
        textfile.write(line)

    textfile.write(str(round(kccd_list[-1],2)) + '@' + str(round(kcl_list[-1],2)) + '@' + str(round(float(ws.cell(maxRow,4).value.replace(",",".")),2)) + '\n')
    textfile.close()

def processXLSX2(path:str, maxRow:int, even:bool):
    wb =  load_workbook(path)
    kcl_list = [0] 
    kccd_list = [0]
    ws = wb.active

    kcl:float = 0
    filename = path.split('//')[-1]
    textfile = open(filename + '.txt',mode='w') 
    count = 2;
    if not even:
        for i in range(2,maxRow+1):
            kcl += sqrt((ws.cell(i,2).value - ws.cell(i-1,2).value)**2 + (ws.cell(i,3).value - ws.cell(i-1,3).value)**2)
            if i%2 == 1:
                kcl_list.append(kcl)
                kccd = float(kccd_list[count-2]) + float(kcl_list[count-1])
                kccd_list.append(kccd)

                kcl = 0
                line:str = str(round(kccd_list[count-2],2)) + '@' + str(round(kcl_list[count-2],2)) + '@' + str(round(float(ws.cell(i-1,4).value),2)) + '\n'
                textfile.write(line)
                print(count)
                count += 1;
        
        textfile.write(str(round(kccd_list[-1],2)) + '@' + str(round(kcl_list[-1],2)) + '@' + str(round(float(ws.cell(maxRow,4).value),2)) + '\n')
                
    else:
        for i in range(2,maxRow+1):
            kcl += sqrt((ws.cell(i,2).value - ws.cell(i-1,2).value)**2 + (ws.cell(i,3).value - ws.cell(i-1,3).value)**2)
            if i%2 == 0:
                kcl_list.append(kcl)
                kccd = float(kccd_list[count-2]) + float(kcl_list[count-1])
                kccd_list.append(kccd)

                kcl = 0
                line:str = str(round(kccd_list[count-2],2)) + '@' + str(round(kcl_list[count-2],2)) + '@' + str(round(float(ws.cell(i-1,4).value),2)) + '\n'
                textfile.write(line)
                print(count)
                count += 1;

        textfile.write(str(round(kccd_list[-1],2)) + '@' + str(round(kcl_list[-1],2)) + '@' + str(round(float(ws.cell(maxRow,4).value),2)) + '\n')


    textfile.close()        

def copyXLSX(path:str, maxRow:int, even:bool, cols:list):
    wb = load_workbook(path)
    ws = wb.active

    count = 1

    wb2 = Workbook()
    ws2 = wb2.active

    for i in range(even+1, maxRow + 1,2):
        for col in cols:
            try:
                c = int(col)
            except:
                print("Not valid column index (not a number)")
                exit(0)
            ws2.cell(count,c).value = ws.cell(i,c).value

        ws2.cell(count,int(cols[-1])).value = 'c' + str(count)

        count += 1
    
    wb2.save(path[:-5] + ' copy.xlsx')

def calculateDistance(path:str, maxRow:int): #process xlsx file contains coordinates
    wb = load_workbook(path)
    ws = wb.active

    for i in range(1,maxRow):
        ws.cell(i,6).value = sqrt(1/100 + abs(float(ws.cell(i,4).value.replace(",",".")) - float(ws.cell(i+1,4).value.replace(",",".")))**2)

    wb.save(path)    

def xlsxToTextFile(path:str,maxRow:int):
    wb = load_workbook(path)
    ws = wb.active

    filename = path.split('//')[-1]
    textfile = open(filename+'.txt',mode='w')

    for i in range(1,maxRow):
        text = str(round(ws.cell(i,3).value,2)) + "@" + str(round(ws.cell(i,2).value,2)) + "@" + str(ws.cell(i,4).value)
        textfile.write(text)

    textfile.close()

def stackDistance(path:str):
        saveFilePath = getPath(path) + "stackDistance.txt"

        with open(path,mode='r') as file:
            lines = file.readlines()
            distance = 0
            stackDistance = 0
            for i in range(1,len(lines)):
                pointA = Point(
                    x=float(lines[i-1].split('@')[0]),
                    y=float(lines[i-1].split('@')[1]),
                    z=float(lines[i-1].split('@')[2])
                )
                pointB = Point(
                    x=float(lines[i].split('@')[0]),
                    y=float(lines[i].split('@')[1]),
                    z=float(lines[i].split('@')[2])
                )
                distance += Point.distance(pointA,pointB)
                if distance > 7:
                    stackDistance += distance
                    info = f'{round(stackDistance,2)}@{round(distance,2)}@{round(pointB.z,2)}\n'
                    with open(saveFilePath,mode='a') as f:
                        f.write(info)
                    distance = 0

def calculatedXLSXtoTXT(path:str, maxRow:int):
    wb = load_workbook(path)
    ws = wb.active
    stackDistance = 0
    stackStackDistance = 0
    textfilePath = getPath(path) + 'cal-ed.txt'
    count = 0
    kclb = kclbb = 0
    for i in range(1,maxRow):
        
        caodo = ws.cell(i,1).value
        khoangcachle = ws.cell(i,2).value
        kclb += 0.1
        stackDistance += khoangcachle
        if (stackDistance > 6.9):
            kclbb += kclb
            if stackDistance > 7.1:
                count += 1
            stackStackDistance += stackDistance
            info = f'{round(kclbb,2)}@{round(kclb,2)}@{round(caodo,2)}\n'
            stackDistance = 0
            kclb = 0
            with open(textfilePath,'a') as textfile:
                textfile.write(info)
    print(f'there are {count} overloaded')

def xlsxToTextFile(path:str, maxRow:int):
    wb = load_workbook(path)
    ws = wb.active

    savePath = getPath(path) + "cal-ed.txt"
    for i in range(1,maxRow+1):
        info = f'{round(ws.cell(i,1).value,2)}@{round(ws.cell(i,2).value,2)}@{round(ws.cell(i,3).value,2)}\n'
        with open(savePath,'a') as f:
            f.write(info)

if __name__ == "__main__":
    if len(cmdArgs) < 3:
        # xlsxToTextFile("G:\\Code\\Python\\autoCAD\\text-file\\cal-ed.xlsx",120)
        # textFileToXLSX("G:\\Code\\Python\\autoCAD\\text-file\\TEST.txt")
        # calculateDistance("G:\\Code\\Python\\autoCAD\\text-file\\text.xlsx",120)
        # processXLSX("G:\\Code\\Python\\autoCAD\\text-file\\text.xlsx",120)
        # stackDistance("G:\\Code\\Python\\autoCAD\\text-file\\TEST.txt")
        # textFileToXLSX("G:\\Code\\Python\\autoCAD\\text-file\\cal-ed.txt")
        # calculatedXLSXtoTXT("G:\\Code\\Python\\autoCAD\\text-file\\test-1.xlsx",5952)
        xlsxToTextFile("G:\\Code\\Python\\autoCAD\\text-file\\test-1.xlsx",99)
        # print("Not enough arguments!")
        # print("__filename.py__ __option__ __number of rows__")
        exit(0)

    file:str = input("Enter dir to input file: ")

    if cmdArgs[1] == 'even': #lẻ
        print(f'processing {file} with option even, there are {cmdArgs[2]} rows')
        processXLSX2(file, int(cmdArgs[2]), True)
    elif cmdArgs[1] == 'odd': #chẵn
        print(f'processing {file} with option odd, there are {cmdArgs[2]} rows')
        processXLSX2(file, int(cmdArgs[2]), False)
    elif cmdArgs[1] == 'all': #tất cả
        print(f'processing {file} with option all, there are {cmdArgs[2]} rows')
        processXLSX(file,int(cmdArgs[2]))
    elif cmdArgs[1] == 'copy':
        copyOpt = input("Copy option: ")
        columnOpt = input("Copy columns: ")

        columnList:list
        if columnOpt == "d": #default 
            columnList = [2, 3, 4, 5]
        else:
            columnList = columnOpt.split(' ')

        if copyOpt == 'even': #chẵn 
            copyXLSX(file, int(cmdArgs[2]), True, columnList)
        elif copyOpt == 'odd': #lẻ
            copyXLSX(file, int(cmdArgs[2]), False, columnList)
        else: 
            print("Not valid copy option, retry")
            exit(0)
    else:
        print("Not valid option, retry")
        exit(0)